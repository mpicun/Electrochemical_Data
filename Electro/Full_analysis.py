#!/usr/bin/env python
# coding: utf-8

import os 
import sys
from math import log10, floor, ceil
import numpy as np
from scipy import integrate
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from matplotlib.ticker import FormatStrFormatter
from openpyxl import Workbook, load_workbook
import xlsxwriter

#Input parameters
path_to_input = sys.argv[1]
path_to = os.path.splitext(path_to_input)[0]
v_thickness = os.path.basename(os.path.dirname(os.path.dirname(path_to_input))) 
sample_name = os.path.basename(path_to).split("_")[0]+"_VACNT_"+str(v_thickness)+"nmV_400°C"

#V2O5 load from deposition of metallic vanadium in nanometers (quartz u-balance readout)
mm_v = 50.94150                                            # g/mol molar mass of o
mm_o = 15.99940                                            # g/mol molar mass of v
mm_v2o5 = 2*mm_v+5*mm_o                                    # g/mol molar mass of v2o5          
v_density = 6110                                           # mg/cm^3 vanadium metal density
elec_area = (np.pi)*(0.615**2)                             # electrode area in cm^2
v_load = float(v_thickness)*elec_area*v_density*10**(-7)   # v metal mass load in mg
tooling = 3/3.2                                            # tooling factor for e-beam dep.
v2o5_load = v_load*(mm_v2o5/(2*mm_v))*tooling              # v2o5 mass load in mg

#Internal function to derivate
def derivate(V, C, dV_step):
    dCdV = []
    Voltage = []
    
    V_last = V[0]
    C_last = C[0]
    
    for i in range(1, len(V)):
        delta_V = V[i] - V_last        
        if abs(delta_V) > dV_step:
            delta_C = C[i] - C_last  
            derivative = delta_C / delta_V
            dCdV.append(derivative)
            Voltage.append(V[i])
            C_last = C[i]
            V_last = V[i]
    return  Voltage, dCdV

#Open data file and get the current unit (uA or mA)
data_file = load_workbook(filename=path_to_input)
t_I_V = data_file["Hoja1"]
N_row_max = t_I_V.max_row

I_unit=t_I_V[1][1].value.split('/')[1]
if I_unit == 'mA': I_unit = 1000
elif I_unit == 'uA': I_unit = 1

t_unit=t_I_V[1][0].value.split('/')[1]
if t_unit == 'ms': t_unit=10**-3
elif t_unit == 's': t_unit=1
    
I_unit, t_unit

#First analysis: Saving end_points (rows number) an C_rate for each cycle
Discharge_EndPoints = []; Charge_EndPoints = [] 
C_rates = []; I = []
EndPoints = [2]; last_I = 0

for row in t_I_V.iter_rows(min_row=2, max_row=N_row_max):
    N_row = row[0].row
    I_n = row[1].value                                                  # Current in mAh/g 
    
    if last_I*I_n < 0:                                                  # Cycles are performed in the 2-4V range
        average_I = abs(sum(I)/len(I))
        c_rate = (average_I*I_unit)/(v2o5_load*294.0)                       # Current in C_rate units
        c_rate = round(c_rate, -int(floor(log10(c_rate))))
        C_rates.append(c_rate)                                          # Write C-rate with 1 significative in mAh/g
        EndPoints.append(N_row-1)
        if last_I < 0: Discharge_EndPoints.append(EndPoints)
        elif last_I > 0: Charge_EndPoints.append(EndPoints)
        EndPoints = [N_row]
        I = []
    I.append(I_n)
    last_I = I_n

EndPoints.append(N_row_max)
if last_I < 0: Discharge_EndPoints.append(EndPoints)
elif last_I > 0: Charge_EndPoints.append(EndPoints)

C_rates.append(C_rates[-1]); C_rates = C_rates[1::2]
C_rates = [ int(c_rate) if c_rate >= 1.0 else c_rate for c_rate in C_rates ]

#First analysis: Integrate the current to calculate the capacity
Charge_Voltage  = []
Charge_Capacity  = []                   
                  
Discharge_Voltage  = []
Discharge_Capacity  = [] 

# Extract the selected Discharge cycle
for EndPoints in Discharge_EndPoints:
    start, end = EndPoints                 
    t = []; I = []; V = []; Q = []                             # Init (t, I, V, Q) for next cyle
    for row in t_I_V.iter_rows(min_row=start, max_row=end):  
        t_n = (row[0].value)*(t_unit/3600.0)                   # Time in hrs 
        I_n = abs(row[1].value)*(I_unit/v2o5_load)             # Current in mA/g
        V_n = (row[2].value)                                   # Discharge Potential, in V vs Li⁺/Li
            
        t.append(t_n)
        I.append(I_n)
        V.append(V_n)

    Q = integrate.cumulative_trapezoid(I, t, initial=0)       # Cumulative Discharge Capacity in mAh/g
    Discharge_Voltage.append(V)
    Discharge_Capacity.append(Q)

# Extract the selected Charge cycle
for EndPoints in Charge_EndPoints:
    start, end = EndPoints    
    t = []; I = []; V = []; Q = []                           # Init (t, I, V, Q) for next cyle
    for row in t_I_V.iter_rows(min_row=start, max_row=end):  
        t_n = (row[0].value)*(t_unit/3600.0)                   # Time in hrs 
        I_n = abs(row[1].value)*(I_unit/v2o5_load)             # Current in mA/g
        V_n = (row[2].value)                                   # Charge Potential, in V vs Li⁺/Li
            
        t.append(t_n)
        I.append(I_n)
        V.append(V_n)
        
    Q = integrate.cumulative_trapezoid(I, t, initial=0)      # Cumulative Charge Capacity in mAh/g
    Charge_Voltage.append(V)
    Charge_Capacity.append(Q)
    
#Saving the Charge/Discharge curves
path_to_output_1 = path_to+"_cycles.xlsx"
workbook_1 = xlsxwriter.Workbook(path_to_output_1)
ws = workbook_1.add_worksheet()

ws.write(0, 0, "Capacidad(mAh/g)")

last_nrow = 1
for n_cycle in range(len(Discharge_Capacity)):
    #Writing n-Discharge curve
    ws.write(0, 2*n_cycle+1,"Descarga_"+str(n_cycle+1)+" (V)")
    for n_row in range(len(Discharge_Capacity[n_cycle])):
        ws.write(n_row+last_nrow, 0, Discharge_Voltage[n_cycle][n_row])
        ws.write(n_row+last_nrow, 2*n_cycle+1, Discharge_Capacity[n_cycle][n_row])        
    last_nrow += len(Discharge_Capacity[n_cycle])
    #Writing n-Charge curve
    ws.write(0, 2*n_cycle+2,"Carga_"+str(n_cycle+1)+" (V)")
    for n_row in range(len(Charge_Capacity[n_cycle])):
        ws.write(n_row+last_nrow, 0, Charge_Voltage[n_cycle][n_row])
        ws.write(n_row+last_nrow, 2*n_cycle+2, Charge_Capacity[n_cycle][n_row])        
    last_nrow += len(Charge_Capacity[n_cycle])
workbook_1.close()

#Second analysis: extract the C-rates capacity
Discharge_Capacity_N=[]; Charge_Capacity_N=[]

for cycle in range(len(Discharge_Capacity)): Discharge_Capacity_N.append(Discharge_Capacity[cycle][-1])
for cycle in range(len(Charge_Capacity)): Charge_Capacity_N.append(Charge_Capacity[cycle][-1])

Max_N_cycles = np.min([len(Discharge_Capacity_N), len(Charge_Capacity_N)])

#Saving the C-rates capacity
path_to_output_2 = path_to+"_Crates.xlsx"
workbook_2 = xlsxwriter.Workbook(path_to_output_2)
ws = workbook_2.add_worksheet()

ws.write(0, 0, "Número de Ciclo")
ws.write(0, 1, "Capacidad de Descarga(mAh/g)")    
ws.write(0, 2, "Capacidad de Carga(mAh/g)")
ws.write(0, 3, "C-rate")

for cycle_number in range(Max_N_cycles):
    ws.write(cycle_number+1, 0, cycle_number+1)
    ws.write(cycle_number+1, 1, Discharge_Capacity_N[cycle_number])
    ws.write(cycle_number+1, 2, Charge_Capacity_N[cycle_number])
    ws.write(cycle_number+1, 3, C_rates[cycle_number])

workbook_2.close()

#Third analysis: Derivate the Capacity to obtain dQ/dV curve for first cycles
ID_analized_cycles = [0,1,2,3,4]
dV_step = 0.005

Voltage_Discharge_Phase = []
Voltage_Charge_Phase = []
dQdV_Discharge_Phase = []
dQdV_Charge_Phase = []

path_to_output_3 = path_to+"_dQdV.xlsx"
workbook_3 = xlsxwriter.Workbook(path_to_output_3)
ws1 = workbook_3.add_worksheet()

ws1.write(0, 0, "Potencial (V vs Li+/Li)")
n_col = 1
last_nrow = 1

for ID in ID_analized_cycles:
    ws1.write(0, n_col, "dQdV descarga_"+str(ID))
    
    Discharge_Voltage_ID, dQdV_discharge_ID  = derivate(Discharge_Voltage[ID], Discharge_Capacity[ID], dV_step)
    Voltage_Discharge_Phase.append(Discharge_Voltage_ID)
    dQdV_Discharge_Phase.append(dQdV_discharge_ID)
    
    for n_row in range(len(dQdV_discharge_ID)): 
        ws1.write(last_nrow + n_row, 0, Discharge_Voltage_ID[n_row])
        ws1.write(last_nrow + n_row, n_col, dQdV_discharge_ID[n_row])
    n_col += 1
    last_nrow += len(dQdV_discharge_ID)
    
    Charge_Voltage_ID, dQdV_charge_ID  = derivate(Charge_Voltage[ID], Charge_Capacity[ID], dV_step)  
    Voltage_Charge_Phase.append(Charge_Voltage_ID)
    dQdV_Charge_Phase.append(dQdV_charge_ID)
  
    ws1.write(0, n_col, "dQdV carga_"+str(ID))
    for n_row in range(len(dQdV_charge_ID)): 
        ws1.write(last_nrow + n_row, 0, Charge_Voltage_ID[n_row])
        ws1.write(last_nrow + n_row, n_col, dQdV_charge_ID[n_row])
    n_col += 1
    last_nrow += len(dQdV_charge_ID)        

ws2 = workbook_3.add_worksheet()

ws2.write(0, 0, "Capacidad(mAh/g)")

last_nrow = 1
for n_cycle in ID_analized_cycles:
    #Writing n-Discharge curve
    ws2.write(0, 2*n_cycle+1,"Descarga_"+str(n_cycle+1)+" (V)")
    for n_row in range(len(Discharge_Capacity[n_cycle])):
        ws2.write(n_row+last_nrow, 0, Discharge_Voltage[n_cycle][n_row])
        ws2.write(n_row+last_nrow, 2*n_cycle+1, Discharge_Capacity[n_cycle][n_row])        
    last_nrow += len(Discharge_Capacity[n_cycle])
    #Writing n-Charge curve
    ws2.write(0, 2*n_cycle+2,"Carga_"+str(n_cycle+1)+" (V)")
    for n_row in range(len(Charge_Capacity[n_cycle])):
        ws2.write(n_row+last_nrow, 0, Charge_Voltage[n_cycle][n_row])
        ws2.write(n_row+last_nrow, 2*n_cycle+2, Charge_Capacity[n_cycle][n_row])        
    last_nrow += len(Charge_Capacity[n_cycle])
    
workbook_3.close()            

#Plot all Charge/Discharge curves
y_range = 4
x_range = ceil(np.max([Discharge_Capacity_N, Charge_Capacity_N])/100)*100

fig, ax = plt.subplots(figsize=(10, 6), layout='constrained') 
plt.suptitle("Test Galvanostático Carga/Descarga", x=0.54, fontsize=19)

for n_cycle in range(len(Discharge_Capacity)): ax.plot(Discharge_Capacity[n_cycle], Discharge_Voltage[n_cycle], color="indigo", linewidth=1.5, label="Descarga")
for n_cycle in range(len(Charge_Capacity)): ax.plot(Charge_Capacity[n_cycle], Charge_Voltage[n_cycle], color="orangered",  linewidth=1.5, label="Carga")

ax.set_title(sample_name, fontsize=17)
ax.set_xlabel("Capacidad (mAh/g)", fontsize=17)
ax.set_ylabel(r"Potential (V vs Li/Li$^+$)",fontsize=17)

discharge_legend = mlines.Line2D([], [], color='indigo', label='Descarga')
charge_legend = mlines.Line2D([], [], color='orangered', label='Carga')
ax.legend(handles=[charge_legend, discharge_legend], fontsize=16, loc="right")

ax.set_ylim(2, y_range)
ax.set_xlim(0, x_range)
ax.set_yticks(np.arange(2, 4.5, 0.5),)
ax.set_yticks(np.arange(2, y_range, 0.1), minor=True)
ax.set_xticks(np.arange(0, x_range, 10), minor=True)
ax.tick_params(axis='both', which='major', labelsize=14)

ax.grid(which='minor', alpha=0.2)
ax.grid(which='major', alpha=0.5)

ax.xaxis.set_major_formatter(FormatStrFormatter('%.0f'))

#Save the plot
path_to_jpg = path_to+"_cycles.jpg"     
plt.savefig(path_to_jpg, dpi=450)

#Plot first 5 Charge/Discharge curves for dQ/dV analysis
y_range = 4
x_range = np.max([Discharge_Capacity_N, Charge_Capacity_N])+10

fig, ax = plt.subplots(figsize=(10, 6), layout='constrained') 
plt.suptitle("Test Galvanostático Carga/Descarga", x=0.54, fontsize=19)

for n_cycle in ID_analized_cycles: 
    ax.plot(Discharge_Capacity[n_cycle], Discharge_Voltage[n_cycle], color="indigo", linewidth=1.5)
    ax.plot(Charge_Capacity[n_cycle], Charge_Voltage[n_cycle], color="orangered",  linewidth=1.5)

ax.set_title(sample_name, fontsize=17)
ax.set_xlabel("Capacidad (mAh/g)", fontsize=17)
ax.set_ylabel(r"Potential (V vs Li/Li$^+$)",fontsize=17)

discharge_legend = mlines.Line2D([], [], color='indigo', label='Descarga')
charge_legend = mlines.Line2D([], [], color='orangered', label='Carga')
ax.legend(handles=[charge_legend, discharge_legend], fontsize=16, loc="right")

ax.set_ylim(2, y_range)
ax.set_xlim(0, x_range)
ax.set_yticks(np.arange(2, 4.5, 0.5),)
ax.set_yticks(np.arange(2, y_range, 0.1), minor=True)
ax.set_xticks(np.arange(0, x_range, 10), minor=True)
ax.tick_params(axis='both', which='major', labelsize=14)

ax.grid(which='minor', alpha=0.2)
ax.grid(which='major', alpha=0.5)

ax.xaxis.set_major_formatter(FormatStrFormatter('%.0f'))

#Save the plot
path_to_jpg = path_to+"_5first.jpg"     
plt.savefig(path_to_jpg, dpi=450)

#Plot the C-rates
y_range = ceil(np.max([Discharge_Capacity_N, Charge_Capacity_N])/100)*100
x_range = len(Discharge_Capacity_N)

fig, ax = plt.subplots(figsize=(10, 6), layout='constrained') 
plt.suptitle("Test Galvanostático Carga/Descarga", x=0.54, fontsize=19)

cycle_number = np.arange(1,len(Charge_Capacity_N)+1)
ax.plot(cycle_number, Charge_Capacity_N, color="orangered", marker="^", linestyle='None', markersize=9, label='Carga')
ax.plot(cycle_number, Discharge_Capacity_N, color="indigo", marker="v", linestyle='None', markersize=9, label='Descarga')

ax.set_title(sample_name, fontsize=17)
ax.set_xlabel("Número de ciclo", fontsize=17)
ax.set_ylabel("Capacidad (mAh/g)", fontsize=17)

ax.set_ylim(0, y_range)
ax.set_xlim(0, x_range+1)
ax.set_yticks(np.arange(0, y_range+50, 50),)
ax.set_yticks(np.arange(0, y_range, 10), minor=True)
ax.tick_params(axis='both', which='major', labelsize=14)

ax.grid(which='minor', alpha=0.2)
ax.grid(which='major', alpha=0.5)

ax.xaxis.set_major_formatter(FormatStrFormatter('%.0f'))

#Centering c-rate labels
n_cycle, pos, count = 0, 0, 0
last_C_r = C_rates[0]

for C_r in C_rates:
    n_cycle += 1
    if C_r != last_C_r or n_cycle == len(C_rates): 
        pos = floor(n_cycle-count/2)
        ax.annotate(str(last_C_r)+"C", (pos, Discharge_Capacity_N[pos]+25), fontsize=14)         
        count = 0
    count += 1
    last_C_r = C_r 

ax.legend(fontsize=17)

#Save the plot
path_to_jpg = path_to+"_Crates.jpg"     
plt.savefig(path_to_jpg, dpi=450)

#Plot the dQ/dV curves
fig, ax = plt.subplots(figsize=(10, 6), layout='constrained') 
plt.suptitle("Análisis de la Capacidad Diferencial", x=0.54, fontsize=19)

ax.set_title(sample_name, fontsize=17)
ax.set_xlabel(r"Potencial (V vs Li/Li$^+$)", fontsize=17)
ax.set_ylabel("dQ/dV (mAh/gV)",fontsize=17)

for n_cycle in range(len(ID_analized_cycles)): 
    ax.plot(Voltage_Discharge_Phase[n_cycle], dQdV_Discharge_Phase[n_cycle], color="indigo", linewidth=1.5)
    ax.plot(Voltage_Charge_Phase[n_cycle], dQdV_Charge_Phase[n_cycle], color="orangered",  linewidth=1.5)


ax.set_xlim(2, 4)
ax.set_xticks(np.arange(2, 4, 0.1), minor=True)
ax.tick_params(axis='both', which='major', labelsize=14)

ax.grid(which='minor', alpha=0.2)
ax.grid(which='major', alpha=0.5)

ax.xaxis.set_major_formatter(FormatStrFormatter('%.1f'))

discharge_legend = mlines.Line2D([], [], color='indigo', label='Descarga')
charge_legend = mlines.Line2D([], [], color='orangered', label='Carga')

ax.legend(handles=[charge_legend, discharge_legend], fontsize=16, loc="lower right")

#Save the plot
path_to_jpg = path_to+"_phases.jpg"     
plt.savefig(path_to_jpg, dpi=450)
